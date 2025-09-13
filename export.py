import pandas as pd
from openpyxl import load_workbook
from openpyxl.cell.cell import MergedCell
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
from tkinter import Tk, Label, Button, filedialog, messagebox
import os
import subprocess
import sys

# --- Tùy chọn dấu phân tách trong công thức Excel ---
USE_SEMICOLON = True
SEP = ";" if USE_SEMICOLON else ","

# --- Hàm lấy đường dẫn file tài nguyên (template.xlsx) ---
def resource_path(relative_path):
    try:
        # PyInstaller tạo thư mục tạm _MEIPASS
        base_path = sys._MEIPASS
    except Exception:
        base_path = os.path.abspath(".")
    return os.path.join(base_path, relative_path)

# --- Hàm tạo báo cáo ---
def generate_report(source_file, target_file):
    template_file = resource_path("template.xlsx")
    if not os.path.exists(template_file):
        raise FileNotFoundError(f"Không tìm thấy template.xlsx tại {template_file}")

    df = pd.read_excel(source_file, header=0)
    df = df.iloc[:, [5, 6, 7, 8, 9, 11]]
    df.columns = ["Mã nhân viên", "Tên nhân viên", "Vị trí", "Col_I", "Col_J", "Col_L"]

    order = pd.CategoricalDtype(categories=["SPV", "Admin", "Shipper-chính thức"], ordered=True)
    df["Vị trí"] = df["Vị trí"].astype(order)
    df_sorted = df.sort_values(by=["Vị trí", "Tên nhân viên"])

    wb = load_workbook(template_file)
    ws = wb.active
    start_row = 5

    # Xóa dữ liệu cũ
    for row in ws.iter_rows(min_row=start_row, max_row=ws.max_row, min_col=1, max_col=8):
        for cell in row:
            if not isinstance(cell, MergedCell):
                cell.value = None

    # Ghi dữ liệu mới
    for r_idx, row in enumerate(df_sorted.itertuples(index=False), start=start_row):
        ws.cell(row=r_idx, column=1, value=row[2])
        ws.cell(row=r_idx, column=2, value=row[0])
        ws.cell(row=r_idx, column=3, value=row[1])
        ws.cell(row=r_idx, column=5, value=row[3])
        ws.cell(row=r_idx, column=6, value=row[4])
        ws.cell(row=r_idx, column=8, value=row[5])

        formula = f"=IF(OR((H{r_idx}+E{r_idx})=0,(H{r_idx}+F{r_idx})/(H{r_idx}+E{r_idx})>1),\"\",(H{r_idx}+F{r_idx})/(H{r_idx}+E{r_idx}))"
        cell_G = ws.cell(row=r_idx, column=7, value=formula)
        cell_G.number_format = "0.00%"

    # Merge và định dạng các ô
    current_value = None
    merge_start = start_row
    for r in range(start_row, start_row + len(df_sorted)):
        value = ws.cell(row=r, column=1).value
        if value != current_value:
            if current_value is not None and r - merge_start > 1:
                ws.merge_cells(start_row=merge_start, start_column=1, end_row=r - 1, end_column=1)
            current_value = value
            merge_start = r
    if current_value is not None and (start_row + len(df_sorted) - merge_start) > 1:
        ws.merge_cells(start_row=merge_start, start_column=1, end_row=start_row + len(df_sorted) - 1, end_column=1)

    end_row = start_row + len(df_sorted) - 1
    if end_row >= start_row:
        ws.merge_cells(start_row=start_row, start_column=4, end_row=end_row, end_column=4)
        for col in range(9, 17):
            ws.merge_cells(start_row=start_row, start_column=col, end_row=end_row, end_column=col)
        ws.merge_cells(start_row=start_row, start_column=17, end_row=end_row, end_column=19)

    alignment = Alignment(horizontal="center", vertical="center")
    font_normal = Font(name="Times New Roman", size=11, bold=False)
    font_bold = Font(name="Times New Roman", size=11, bold=True)
    thin = Side(border_style="thin", color="000000")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    for row in ws.iter_rows(min_row=start_row, max_row=end_row, min_col=1, max_col=8):
        for cell in row:
            cell.alignment = alignment
            cell.border = border
            cell.font = font_bold if cell.column == 1 else font_normal
    for row in ws.iter_rows(min_row=start_row, max_row=end_row, min_col=9, max_col=19):
        for cell in row:
            cell.alignment = alignment
            cell.border = border
            cell.font = font_normal

    # Hàng tổng
    total_row = end_row + 1
    ws.merge_cells(start_row=total_row, start_column=1, end_row=total_row, end_column=3)
    cell_total = ws.cell(row=total_row, column=1, value="Tổng")

    ws.cell(row=total_row, column=4, value=f"=D{start_row}")
    ws.cell(row=total_row, column=5, value=f"=SUM(E{start_row}:E{end_row})")
    ws.cell(row=total_row, column=6, value=f"=SUM(F{start_row}:F{end_row})")
    ws.cell(row=total_row, column=8, value=f"=SUM(H{start_row}:H{end_row})")

    cell_g_total = ws.cell(row=total_row, column=7, value=f"=AVERAGE(G{start_row}:G{end_row})")
    cell_g_total.number_format = "0.00%"

    for col in range(9, 17):
        col_letter = chr(64 + col)
        ws.cell(row=total_row, column=col, value=f"={col_letter}{start_row}")

    ws.merge_cells(start_row=total_row, start_column=17, end_row=total_row, end_column=19)
    ws.cell(row=total_row, column=17, value=None)

    header_fill = ws[1][0].fill  # fill của ô A1 (hàng đầu tiên)

    # Tạo một PatternFill mới dựa trên header_fill
    copied_fill = PatternFill(
        fill_type=header_fill.fill_type,
        start_color=header_fill.start_color,
        end_color=header_fill.end_color
        ) 

    # Định dạng toàn bộ A..S của hàng tổng
    for cell in ws[total_row][0:19]:  # A..S
        cell.font = font_bold
        cell.alignment = alignment
        cell.border = border
        cell.fill = copied_fill

    # Công thức J5
    ws.cell(row=start_row, column=10, value=f"=I{start_row}/($F${total_row}+$H${total_row})").number_format = "0.00%"

    # Công thức M5 (chia tổng F cho số lượng nhân viên)
    num_employees = end_row - start_row + 1
    ws.cell(row=start_row, column=13, value=f"=$F${total_row}/{num_employees}")

    # Lưu file
    wb.save(target_file)
    os.startfile(target_file)
    return target_file

# --- UI Tkinter ---
def run_ui():
    root = Tk()
    root.title("Tạo báo cáo vận hành")
    root.geometry("500x200")

    # Lấy thư mục EXE chạy
    if getattr(sys, 'frozen', False):
        exe_dir = os.path.dirname(sys.executable)
    else:
        exe_dir = os.path.dirname(os.path.abspath(__file__))

    def choose_source():
        file = filedialog.askopenfilename(filetypes=[("Excel files", "*.xlsx")])
        if file:
            source_label.config(text=file)
            root.source_file = file

    def run_process():
        try:
            source_file = getattr(root, "source_file", None)
            if not source_file:
                messagebox.showerror("Lỗi", "Hãy chọn file dữ liệu.")
                return
            target_file = os.path.join(exe_dir, "BÁO CÁO VẬN HÀNH BƯU CỤC.xlsx")
            result = generate_report(source_file, target_file)
            root.target_file = result
            messagebox.showinfo("Thành công", f"Đã tạo báo cáo:\n{result}\nFile đã được mở.")
        except Exception as e:
            messagebox.showerror("Lỗi", str(e))

    def open_file_location():
        try:
            subprocess.Popen(f'explorer "{exe_dir}"')
        except Exception:
            os.startfile(exe_dir)

    Label(root, text="Chọn file dữ liệu:").pack()
    Button(root, text="Browse...", command=choose_source).pack()
    source_label = Label(root, text="Chưa chọn")
    source_label.pack()

    Button(root, text="Chạy tạo báo cáo", command=run_process).pack(pady=10)
    Button(root, text="Mở thư mục chứa file báo cáo", command=open_file_location).pack()

    root.mainloop()


if __name__ == "__main__":
    run_ui()
